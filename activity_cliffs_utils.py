# ================================================================
# Activity Cliffs Calculator - Utility Functions
# Author: Santi Garcia-Vallvé
# ================================================================

import pandas as pd
from itertools import combinations
from tqdm import tqdm
from io import BytesIO

# RDKit imports
from rdkit import Chem, DataStructs
from rdkit.Chem import AllChem, Draw, RDKFingerprint, MACCSkeys, rdFingerprintGenerator, rdMolDescriptors

# Excel handling
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Display utilities
from IPython.display import display, HTML

# Definitions of some functions

# ---------- Fingerprint and pair generation functions ----------

def fp_as_bitvect(mol, fp_type="feature_morgan", n_bits=2048, radius=2, **kwargs):
    """
    Modern RDKit fingerprint generator (no deprecation warnings)
    Compatible with RDKit 2025.01+.
    Supported types: morgan, feature_morgan, rdk, maccs, pattern, torsion, atompair
    """
    if mol is None:
        return None

    fp_type = fp_type.lower().strip()

    if fp_type == "morgan":
        gen = rdFingerprintGenerator.GetMorganGenerator(
            radius=radius, fpSize=n_bits, countSimulation=False
        )
        return gen.GetFingerprint(mol)

    elif fp_type in ("feature_morgan", "fcfp"):
        # RDKit ≥ 2025.03
        if hasattr(rdFingerprintGenerator, "GetMorganFeatureGenerator"):
            fgen = rdFingerprintGenerator.GetMorganFeatureGenerator(
                radius=radius, fpSize=n_bits, countSimulation=False
            )
            return fgen.GetFingerprint(mol)
        # Older versions
        elif hasattr(rdFingerprintGenerator, "GetMorganFeatureAtomInvGen"):
            invgen = rdFingerprintGenerator.GetMorganFeatureAtomInvGen()
            gen = rdFingerprintGenerator.GetMorganGenerator(
                radius=radius, fpSize=n_bits, countSimulation=False,
                atomInvariantsGenerator=invgen
            )
            return gen.GetFingerprint(mol)
        else:
            inv = rdMolDescriptors.GetFeatureInvariants(mol)
            return AllChem.GetMorganFingerprintAsBitVect(
                mol, radius, nBits=n_bits, invariants=inv
            )

    elif fp_type == "rdk":
        return RDKFingerprint(mol, fpSize=n_bits, **kwargs)

    elif fp_type == "maccs":
        return MACCSkeys.GenMACCSKeys(mol)

    elif fp_type == "pattern":
        if hasattr(rdFingerprintGenerator, "GetPatternGenerator"):
            gen = rdFingerprintGenerator.GetPatternGenerator(fpSize=n_bits)
            return gen.GetFingerprint(mol)
        from rdkit.Chem import rdmolops
        return rdmolops.PatternFingerprint(mol, fpSize=n_bits)

    elif fp_type == "torsion":
        gen = rdFingerprintGenerator.GetTopologicalTorsionGenerator(fpSize=n_bits)
        return gen.GetFingerprint(mol)

    elif fp_type == "atompair":
        gen = rdFingerprintGenerator.GetAtomPairGenerator(fpSize=n_bits)
        return gen.GetFingerprint(mol)

    else:
        raise ValueError(f"Unsupported fingerprint type: {fp_type}")

def compute_tanimoto(fp1, fp2):
    """Calculates the Tanimoto similarity between two fingerprints"""
    return DataStructs.TanimotoSimilarity(fp1, fp2)

def generate_pairs(df, group_col="PubMedID", fp_col="fp_feature_morgan", show_progress=False):
    """
    Generates all possible pairs of compounds sharing the same group ID and
    computes Tanimoto similarity and activity disparity using a selected fingerprint column.

    Parameters
    ----------
    df : pandas.DataFrame
        Input dataframe containing at least:
        - group_col (e.g. 'PubMedID')
        - 'Compound'
        - 'pIC50'
        - the fingerprint column (default: 'fp_morgan')
    group_col : str
        Column to group compounds by (default 'PubMedID').
    fp_col : str
        Column name containing the fingerprints to use.
    show_progress : bool
        If True, shows a tqdm progress bar.

    Returns
    -------
    pandas.DataFrame
        DataFrame with:
        group_col, Compound1, Compound2, pIC50_1, pIC50_2, pIC50_diff, Tanimoto, Disparity
    """
    pair_data = []

    iterator = df.groupby(group_col)
    if show_progress:
        iterator = tqdm(iterator, desc=f"Generating pairs ({fp_col})")

    for group_id, group in iterator:
        # skip groups missing fingerprints
        group = group.dropna(subset=[fp_col])
        if len(group) < 2:
            continue

        for row1, row2 in combinations(group.itertuples(index=False), 2):
            fp1, fp2 = getattr(row1, fp_col), getattr(row2, fp_col)
            tanimoto = compute_tanimoto(fp1, fp2)
            pIC50_diff = abs(row1.pIC50 - row2.pIC50)
            disparity = pIC50_diff if tanimoto == 1 else pIC50_diff / (1 - tanimoto)

            pair_data.append({
                group_col: group_id,
                'Fingerprint': fp_col.replace("fp_", ""),
                'Compound1': row1.Compound,
                'SMILES1': row1.standardize_smiles,
                'pIC50_1': row1.pIC50,
                'Compound2': row2.Compound,
                'SMILES2': row2.standardize_smiles,
                'pIC50_2': row2.pIC50,
                'pIC50_diff': pIC50_diff,
                'Tanimoto': tanimoto,
                'Disparity': disparity
            })

    return pd.DataFrame(pair_data)


def export_activity_cliffs_to_excel(filtered_df, output_xlsx, image_size=75):
    """
    Export the filtered DataFrame to an Excel workbook, embedding
    molecule PNG images (Mol1_img and Mol2_img) in dedicated columns
    before each compound name.

    Parameters
    ----------
    filtered_df : pandas.DataFrame
        DataFrame containing result data and Mol1_img / Mol2_img columns.
    output_xlsx : str
        Output Excel file path.
    image_size : int
        Image size in pixels (default: 75).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity_Cliffs"

    # Remove potential SVG columns
    df_to_export = filtered_df.drop(
        columns=[c for c in filtered_df.columns if c.endswith("_svg")],
        errors="ignore"
    )

    # Define column order (Fingerprint after Tanimoto)
    base_cols = [
        'PubMedID',
        'Mol1_img', 'Compound1', 'pIC50_1',
        'Mol2_img', 'Compound2', 'pIC50_2',
        'pIC50_diff', 'Tanimoto', 'Fingerprint', 'Disparity'
    ]

    # Determine available columns
    cols_to_write = [c for c in base_cols if c in df_to_export.columns]
    extra_cols = [c for c in df_to_export.columns if c not in base_cols]
    cols_to_write += extra_cols

    # Prepare DataFrame copy (remove binary image objects)
    df_no_img = df_to_export.copy()
    df_no_img["Mol1_img"] = ""
    df_no_img["Mol2_img"] = ""

    # Write headers and rows
    for r_idx, row in enumerate(dataframe_to_rows(df_no_img[cols_to_write], index=False, header=True), 1):
        ws.append(row)

    # Adjust column widths
    col_widths = {
        "A": 12,   # PubMedID
        "B": max(10, image_size / 7.0),   # Mol1_img
        "C": 26,   # Compound1
        "D": 10,   # pIC50_1
        "E": max(10, image_size / 7.0),   # Mol2_img
        "F": 26,   # Compound2
        "G": 10,   # pIC50_2
        "H": 12,   # pIC50_diff
        "I": 12,   # Tanimoto
        "J": 14,   # Fingerprint
        "K": 12,   # Disparity
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Adjust row height based on image size
    row_height = max(70, int(image_size * 0.75))
    for i in range(2, len(df_to_export) + 2):
        ws.row_dimensions[i].height = row_height

    # Vertically center all text
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    # Insert PNG images into Excel cells
    for i, (img1, img2) in enumerate(zip(df_to_export['Mol1_img'], df_to_export['Mol2_img']), start=2):
        if img1:
            pic1 = XLImage(img1)
            pic1.width = pic1.height = image_size
            ws.add_image(pic1, f"B{i}")
        if img2:
            pic2 = XLImage(img2)
            pic2.width = pic2.height = image_size
            ws.add_image(pic2, f"E{i}")

    wb.save(output_xlsx)
    print(f"✅ Excel file saved with molecule images ({image_size}px): {output_xlsx}")


def mol_to_image_bytes(smiles, image_size=(100, 100)):
    """
    Convert a SMILES string to a PNG image (as BytesIO object)
    for embedding into Excel workbooks.

    Parameters
    ----------
    smiles : str
        Molecule SMILES string.
    image_size : tuple(int, int)
        Width and height in pixels (default: (100, 100)).

    Returns
    -------
    BytesIO or None
        PNG image data stream, ready to use with openpyxl.
    """
    try:
        mol = Chem.MolFromSmiles(smiles)
        if mol:
            img = Draw.MolToImage(mol, size=image_size)
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            return buffer
    except Exception:
        pass
    return None


def smiles_to_svg(smiles, size=(120, 120)):
    """
    Convert a SMILES string to an inline SVG (HTML-compatible)
    for visualization in Jupyter or VS Code notebooks.

    Parameters
    ----------
    smiles : str
        Molecule SMILES string.
    size : tuple(int, int)
        Width and height of the SVG drawing.

    Returns
    -------
    str
        SVG markup as a string (without newlines).
    """
    if not isinstance(smiles, str):
        return ""
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return ""
    try:
        drawer = Draw.MolDraw2DSVG(size[0], size[1])
        drawer.DrawMolecule(mol)
        drawer.FinishDrawing()
        svg = drawer.GetDrawingText()
        return svg.replace("\n", "")
    except Exception:
        return ""

def show_molecule_table(df, max_rows=30, img_size=(120, 120)):
    """
    Display a DataFrame with molecule drawings (SVG) for the first rows.

    - If the DataFrame has more than `max_rows` rows, only the first `max_rows`
      molecules are rendered as SVG (for performance).
    - All rows are still shown in the table.
    - Includes PubMedID column for better traceability.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing columns such as SMILES1, SMILES2, Compound1, Compound2, etc.
    max_rows : int, optional
        Maximum number of rows to draw (default: 30).
    img_size : tuple(int, int), optional
        Molecule drawing size in pixels (default: (120, 120)).
    """
    if df is None or len(df) == 0:
        print("⚠️ Empty DataFrame, nothing to display.")
        return

    df = df.copy().reset_index(drop=True)
    n_rows = len(df)
    shown_rows = min(max_rows, n_rows)

    if n_rows > max_rows:
        print(f"🧪 Displaying {shown_rows} of {n_rows} rows with molecule drawings "
              f"(only first {shown_rows} drawn for performance).")
    else:
        print(f"🧪 Displaying all {n_rows} rows with molecule drawings.")

    # Generate SVGs only for the first `shown_rows` rows
    if "SMILES1" in df.columns:
        df.loc[:shown_rows - 1, "Mol1_svg"] = df.loc[:shown_rows - 1, "SMILES1"].apply(
            lambda s: smiles_to_svg(s, size=img_size)
        )
    if "SMILES2" in df.columns:
        df.loc[:shown_rows - 1, "Mol2_svg"] = df.loc[:shown_rows - 1, "SMILES2"].apply(
            lambda s: smiles_to_svg(s, size=img_size)
        )

    # Build the list of columns to display (PubMedID first)
    cols = []
    if "PubMedID" in df.columns:
        cols.append("PubMedID")

    if "Mol1_svg" in df.columns and "Compound1" in df.columns:
        cols += ["Mol1_svg", "Compound1"]
    if "pIC50_1" in df.columns:
        cols.append("pIC50_1")

    if "Mol2_svg" in df.columns and "Compound2" in df.columns:
        cols += ["Mol2_svg", "Compound2"]
    if "pIC50_2" in df.columns:
        cols.append("pIC50_2")

    for c in ["pIC50_diff", "Tanimoto", "Fingerprint", "Disparity"]:
        if c in df.columns:
            cols.append(c)

    # Create HTML table (embed SVG drawings)
    html_table = df[cols].to_html(escape=False, index=False)
    display(HTML(html_table))

